import os
import re
import openpyxl

def read_from_excel(xlsname):
    h = openpyxl.load_workbook(xlsname)
    for sheet in h.worksheets:
        if sheet.title == 'レビュー指摘事項':
            break

    total = 0
    for i in sheet.iter_rows():
        total += 1

    fnames = {}
    flist  = {}
    current = 4
    while current <= total:
        if sheet['K%s' %current].value in ['対応完了', '完了']:
            line = current - 3
            tmp = sheet['D%s' %current].value.strip().split('\n')
            # 有可能同一个文件是跨行的，这种情况下需要将两行合并为一行
            tmp2 = []
            flag = True
            for i in tmp:
                if not i.strip():
                    continue
                if re.match('[0-9a-zA-Z]{1,}\.[0-9a-zA-Z]{1,}', i) and len(tmp2) >=1 and tmp2[-1].find('.') < 0:
                    tmp2[-1] += '#' + i
                    flag = False
                    print(1, tmp, tmp2, i)
                elif re.match('[0-9a-zA-Z]{1,}\.[0-9a-zA-Z]{1,}', i) and tmp[0].find('.') < 0:
                    tmp2.append(tmp[0] + '#' + i)
                    print(2, tmp, tmp2, i)
                else:
                    tmp2.append(i)
                    flag = True

            fnames[line] = tmp2
            for f in fnames[line]:
                if f not in flist:
                    flist[f] = []
                flist[f].append(line)
        current += 1

    h.close()

    return fnames, flist


def write_to_excel(xlsname, flist):
    data = {}
    for k, v in flist.items():
        for x in v:
            if x not in data:
                data[x] = []
            data[x].append(k)

    h = openpyxl.load_workbook(xlsname)
    for sheet in h.worksheets:
        if sheet.title == 'レビュー指摘事項':
            break

    col = len(sheet.column_dimensions.keys()) + 3
    for k, v in data.items():
        w = []
        for x in v:
            w.append('\\'.join(x.split('/')[2:]))
        sheet.cell(row=k+3, column=col).value = '\n'.join(w)
        #if len(w) >= 2:
        #    print(k, w)

    h.save('new_%s' %xlsname)
    h.close()

    return len(data)

